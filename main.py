"""
  Two dependencies omitted on purpose:
  - pa_queries contains modified SQL queries that get formatted with appropriate dates.
  - barb_dates contains date equivalencies for use with the SQLs
"""

from openpyxl import load_workbook
from barb_dates import dates
import pandas as pd 
import pa_queries
import cx_Oracle
import datetime
import sys
import os
import re

ip = 'xxx.xxx.xx.xxx'
port = port
SID = 'SID'
username = 'username'
password = 'password'

# Output directory.
out_dir = 'Path to output dir'

class Orasights:

	def __init__(self):

		self.dsn_tns = cx_Oracle.makedsn(ip, port, SID)
		self.db = cx_Oracle.connect(username, password, self.dsn_tns)
		self.final_dates = []

	def user_input(self):
		"""
		Take user input date as datetime object find end_date as datetime object.
		Convert from datetime object to string.
		Remove dashes and century for BARB date comparison.

		"""		
		strt_date = datetime.datetime.strptime(input("What is the start date?\nDate Format: YYMMDD\n"), '%y%m%d').date()
		end_date = strt_date + datetime.timedelta(days=6)

		# Convert from datetime object to string.
		strt_date = str(strt_date)
		end_date = str(end_date)

		# Strip century and dashes from date.
		strt_date = re.sub('[-]', '', strt_date[2:])
		end_date = re.sub('[-]', '', end_date[2:])

		# Convert start date to BARB week number.
		if strt_date in dates:
			com_day = dates[strt_date]

		com_week = com_day[:-1]

		self.final_dates = strt_date, end_date, com_week

	def prep_query(self):
		"""
		Inserts appropriate startdate/enddate/barb week into the SQL to be used later.

		"""			
		self.first_q = pa_queries.first.format(self.final_dates[2])
		self.second_q = pa_queries.second.format(self.final_dates[0], self.final_dates[1])
		self.third_q = pa_queries.third.format(self.final_dates[1], self.final_dates[0]) # Enddate is passed first on purpose.
		self.fourth_q = pa_queries.fourth.format(self.final_dates[2])


	def run_query(self):
		""" 
		Executes formatted queries and stores results as panda dataframe object to be exported into Excel.
		Returns error if query cannot be executed.

		"""

		try:
			self.first_df = pd.read_sql(self.first_q, con = self.db)
			self.second_df = pd.read_sql(self.second_q, con = self.db)
			self.third_df = pd.read_sql(self.third_q, con = self.db)
			self.fourth_df = pd.read_sql(self.fourth_q, con = self.db)
			

		except cx_Oracle.DatabaseError as e:
			error, = e.args
			if error.code == 1017:
				print('Please check your credentials.')
			else:
				print('Database connection error: %s'.format(e))
			raise

	def df_to_xl(self):
		""" 
		Contains two helper functions that populate sheets in the appropriate template for East/West.
		Helper functions are used to allow for easier maintanence and reformating.
		
		"""
		def write_to_east():
			""" 
			Writes results from pandas object to corresponding sheets in East template Excel file.

			"""
			file_name = 'Work_Book_{0}_to_{1}.xlsx'.format(self.final_dates[0], self.final_dates[1])
			
			book = load_workbook('Workbook')
			writer = pd.ExcelWriter(file_name, engine='openpyxl') 
			writer.book = book
			writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
		

			self.hsp_df.to_excel(writer, sheet_name='Sheet 1', header=None, index=False, startrow = 1)
			self.res_df.to_excel(writer, sheet_name='Sheet 2', header=None, index=False, startrow = 1)
			self.dpr_df.to_excel(writer, sheet_name='Sheet 3', header=None, index=False, startrow = 1)
			self.him_df.to_excel(writer, sheet_name='Sheet 4', header=None, index=False, startrow = 1)

			os.chdir(out_dir)
			writer.save()

		def write_to_west():
			""" 
			Writes results from pandas object to corresponding sheets in West template Excel file.
			
			"""
			os.chdir(sys.path[0]) #Changes working directory back to the location of the script.
			
			file_name = 'Work_Book_{0}_to_{1}.xlsx'.format(self.final_dates[0], self.final_dates[1])

			book = load_workbook('Workbook')
			writer = pd.ExcelWriter(file_name, engine='openpyxl') 
			writer.book = book
			writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
		

			self.hsp_df.to_excel(writer, sheet_name='Sheet 1', header=None, index=False, startrow = 1)
			self.res_df.to_excel(writer, sheet_name='Sheet 2', header=None, index=False, startrow = 1)
			self.dpr_df.to_excel(writer, sheet_name='Sheet 3', header=None, index=False, startrow = 1)
			self.him_df.to_excel(writer, sheet_name='Sheet 4', header=None, index=False, startrow = 1)

			os.chdir(out_dir)
			writer.save()

		write_to_east()
		write_to_west()


	def disconnect(self):
		"""
		Disconnect from the database. If this fails, for instance
		if the connection instance doesn't exist it doesn't really matter.

		"""
		try:
			self.db.close()
			print("Connection terminated.")
		except cx_Oracle.DatabaseError:
			pass


if __name__ == "__main__":

	p = Orasights()

	try:
		p.user_input()
		p.prep_query()
		p.run_query()
		p.df_to_xl()

	# Disconnect to avoid ORA-00018: Maximum number of sessions exceeded. 
	finally:
		p.disconnect()
		sys.exit()
